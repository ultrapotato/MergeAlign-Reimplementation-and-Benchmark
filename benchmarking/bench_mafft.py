import os
import subprocess
from pathlib import Path
import time
import psutil
from openpyxl import Workbook, load_workbook

def load_processed_entries(excel_file):
    processed_entries = set()
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            input_file, matrix_file = row[0], row[1]
            processed_entries.add((input_file, matrix_file))
    return processed_entries

def run_mafft_alignments(input_fasta, matrices_dir, output_dir, input_filename, excel_sheet, processed_entries, wb, excel_file):
    file_output_dir = os.path.join(output_dir, input_filename)
    Path(file_output_dir).mkdir(parents=True, exist_ok=True)
    
    num_sequences = 0
    for line in open(input_fasta, 'r'):
        if line.startswith('>'):
            num_sequences += 1

    total_time = 0
    total_memory_usage = 0
    for matrix_file in os.listdir(matrices_dir):
        output_path = os.path.join(file_output_dir, f"{matrix_file}.aln")
        
        if (input_filename, matrix_file) in processed_entries or Path(output_path).exists():
            print(f"Alignment {matrix_file} for file {input_filename} already processed. Skipping.")
            continue
        
        print(f"Processing matrix: {matrix_file} for file {input_filename}")
        matrix_path = os.path.join(matrices_dir, matrix_file)

        start_time = time.time()
        process = psutil.Process(os.getpid())
        start_memory = process.memory_info().rss / (1024 * 1024)
        
        cmd = f"mafft --quiet --auto --amino --aamatrix {matrix_path} {input_fasta} > {output_path}"
        
        try:
            subprocess.run(cmd, shell=True, check=True)
            print(f"Completed alignment with matrix {matrix_file} for file {input_filename}")
        except subprocess.CalledProcessError as e:
            print(f"Error with matrix {matrix_file} for file {input_filename}: {e}")
            continue
        
        end_time = time.time()
        end_memory = process.memory_info().rss / (1024 * 1024)

        total_time += end_time - start_time
        total_memory_usage += end_memory - start_memory

    excel_sheet.append([
        input_filename, total_time, total_memory_usage, num_sequences
    ])

    wb.save(excel_file)

def process_all_files_in_directory(input_dir, matrices_dir, output_dir, excel_sheet, processed_entries, wb, excel_file):
    input_files = [f for f in Path(input_dir).glob("*") if f.is_file() and not f.suffix]
    
    for input_file in input_files:
        if input_file.name != ".DS_Store":
            print(f"\nProcessing file: {input_file.name}")
            run_mafft_alignments(str(input_file), matrices_dir, output_dir, input_file.stem, excel_sheet, processed_entries, wb, excel_file)

if __name__ == "__main__":
    INPUT_DIR = "bali_in"
    MATRICES_DIR = "mergealign_matrices_true"
    OUTPUT_DIR = "mafft_alignments"
    EXCEL_FILE = "mafft_speed.xlsx"
    
    processed_entries = load_processed_entries(EXCEL_FILE)
    
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Alignment Metrics"
        sheet.append(["Input File", "Total Time Taken (s)", "Total Memory Usage (MB)", "Number of Sequences"])

    start_time = time.time()
    process_all_files_in_directory(INPUT_DIR, MATRICES_DIR, OUTPUT_DIR, sheet, processed_entries, wb, EXCEL_FILE)
    
    print(f"\nTotal time taken for all alignments: {time.time() - start_time:.2f} seconds")
