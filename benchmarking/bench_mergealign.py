import subprocess
from pathlib import Path
import time
import psutil
from openpyxl import Workbook, load_workbook
import os

def run_mergealign(alignments_dir, output_file, excel_sheet, wb, excel_file):
    if not Path(alignments_dir).exists():
        raise ValueError(f"Alignments directory not found: {alignments_dir}")
    
    start_time = time.time()
    process = psutil.Process(os.getpid())
    start_memory = process.memory_info().rss / (1024 * 1024)
    
    cmd = f"python3 mergealign.py -a {alignments_dir} -f {output_file}"
    try:
        result = subprocess.run(cmd, shell=True, check=True, capture_output=True, text=True)
        print(result.stdout)
        print(f"Successful for {alignments_dir}")
        print(f"Output saved to: {output_file}")
        
    except subprocess.CalledProcessError as e:
        print(f"Error for {alignments_dir}: {e}")
        if e.stdout:
            print("stdout:", e.stdout)
        if e.stderr:
            print("stderr:", e.stderr)
    
    end_time = time.time()
    end_memory = process.memory_info().rss / (1024 * 1024)
    
    total_time = end_time - start_time
    total_memory_usage = end_memory - start_memory
    
    excel_sheet.append([alignments_dir.name, total_time, total_memory_usage])
    wb.save(excel_file)

def check_setup():
    required_files = {
        'mergealign.py': 'MergeAlign implementation file',
        'mafft_alignments': 'Directory containing MAFFT alignments'
    }
    
    missing = []
    for file, description in required_files.items():
        if not Path(file).exists():
            missing.append(f"{description} ({file})")
    
    if missing:
        print("Missing required files/directories:")
        for item in missing:
            print(f"- {item}")
        return False
    return True

def process_alignment_folders(alignments_dir, output_dir, excel_sheet, wb, excel_file):
    alignments_path = Path(alignments_dir)
    if not alignments_path.exists() or not alignments_path.is_dir():
        raise ValueError(f"Invalid alignments directory: {alignments_dir}")
    
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    for subfolder in alignments_path.iterdir():
        if subfolder.is_dir():
            print(f"Processing {subfolder.name}...")
            output_file = output_path / f"merged_{subfolder.name}.fasta"
            run_mergealign(subfolder, output_file, excel_sheet, wb, excel_file)

def load_or_create_excel(excel_file):
    if Path(excel_file).exists():
        wb = load_workbook(excel_file)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "MergeAlign Metrics"
        sheet.append(["Input File", "Total Time Taken (s)", "Total Memory Usage (MB)"])
    return sheet, wb

if __name__ == "__main__":
    ALIGNMENTS_DIR = "mafft_alignments"
    OUTPUT_DIR = "merge_alignments"
    EXCEL_FILE = "merge_speed.xlsx"
    
    print("Checking setup...")
    if not check_setup():
        print("Not all required files are present.")
        exit(1)
    
    excel_sheet, wb = load_or_create_excel(EXCEL_FILE)
    
    print(f"Running MergeAlign for each subfolder and saving outputs to '{OUTPUT_DIR}'...")
    try:
        process_alignment_folders(ALIGNMENTS_DIR, OUTPUT_DIR, excel_sheet, wb, EXCEL_FILE)
    except Exception as e:
        print(f"Error during processing: {e}")
    
    print("Done!")
